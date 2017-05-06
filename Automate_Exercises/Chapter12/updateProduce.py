import openpyxl

# My stupid code
# def get_produce_price(produce_name):
#     produce_price = 0
#     
#     if produce_name == "Celery":
#         produce_price = 1.19
#     if produce_name == "Garlic":
#         produce_price = 3.07
#     if produce_name == "Lemon":
#         produce_price = 1.27
#         
#     return produce_price
# 
# produce_wb = openpyxl.load_workbook("produceSales.xlsx")
# produce_sheet = produce_wb.get_active_sheet()
# 
# updated_count = 0
# for i in range(2, produce_sheet.max_row):
#     if produce_sheet["A" + str(i)].value == "Celery":
#         produce_sheet["B" + str(i)].value = get_produce_price("Celery")
#         updated_count += 1
# 
#     if produce_sheet["A" + str(i)].value == "Garlic":
#         produce_sheet["B" + str(i)].value = get_produce_price("Garlic")
#         updated_count += 1
#         
#     if produce_sheet["A" + str(i)].value == "Lemon":
#         produce_sheet["B" + str(i)].value = get_produce_price("Lemon")
#         updated_count += 1
#         
# produce_wb.save("updatedProduceSales.xlsx")
# print(updated_count)

# Good code
PRICE_UPDATES = {
        "Garlic" : 3.07,
        "Celery" : 1.19,
        "Lemon" : 1.27
    }

produce_wb = openpyxl.load_workbook("produceSales.xlsx")
produce_sheet = produce_wb.get_active_sheet()

updated_count = 0
for row_num in range(2, produce_sheet.max_row):
    produce_name = produce_sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        produce_sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]
        updated_count += 1
        
produce_wb.save("updatedProduceSales.xlsx")
print(updated_count)