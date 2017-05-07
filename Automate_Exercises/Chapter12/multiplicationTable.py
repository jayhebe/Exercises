import openpyxl, sys

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

if len(sys.argv) != 2:
    print("Usage: multiplicationTable.py [number]")
else:
    number = int(sys.argv[1])
    
    for row_num in range(1, number + 1):
        sheet.cell(row=row_num + 1, column=1).value = row_num
    
    for column_num in range(1, number + 1):
        sheet.cell(row=1, column=column_num + 1).value = column_num
        
    for data_row in range(1, sheet.max_row):
        for data_column in range(1, sheet.max_column):
            sheet.cell(row=data_row + 1, column=data_column + 1).value = data_row * data_column
        
wb.save("multiplicationTable.xlsx")